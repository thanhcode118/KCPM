
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── COLOR & STYLE HELPERS ───────────────────────────────────────────────────
DARK_BLUE   = "00008B"
HEADER_BLUE = "000080"
LIGHT_BLUE  = "C6EFCE"
WHITE       = "FFFFFF"
YELLOW_BG   = "FFFF00"
ORANGE_BG   = "FFC000"
GREEN_BG    = "92D050"
GREY_BG     = "D9D9D9"

def hdr_fill(color=HEADER_BLUE):
    return PatternFill("solid", fgColor=color)

def white_font(bold=False, size=9):
    return Font(name="Calibri", color=WHITE, bold=bold, size=size)

def dark_font(bold=False, size=9):
    return Font(name="Calibri", color="000000", bold=bold, size=size)

def thin_border():
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def apply_header_cell(ws, row, col, value, bg=HEADER_BLUE, font_bold=False,
                      align="center", wrap=False, span_cols=None, span_rows=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = hdr_fill(bg)
    cell.font = white_font(bold=font_bold)
    cell.border = thin_border()
    cell.alignment = center_align(wrap) if align == "center" else left_align(wrap)
    if span_cols and span_rows is None:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span_cols - 1)
    elif span_cols and span_rows:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row + span_rows - 1, end_column=col + span_cols - 1)
    return cell

def apply_data_cell(ws, row, col, value="", bg=WHITE, bold=False,
                    align="center", wrap=False, color="000000"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", color=color, bold=bold, size=9)
    cell.border = thin_border()
    cell.alignment = center_align(wrap) if align == "center" else left_align(wrap)
    return cell

# ─── TEST CASE DATA ───────────────────────────────────────────────────────────
# Format: (sheet_name, function_name, created_by, lines_of_code, jira_ticket,
#          test_cases_list)
# Each test_case: (utcid, description, precondition, inputs, expected_return,
#                  exception, log_message, type, result, exec_date)

SHEETS = []

# ══════════════════════════════════════════════════════════════════════════════
# 1. LOGIN  (UserServiceLoginTests)
# ══════════════════════════════════════════════════════════════════════════════
login_tcs = [
    # (UTCID, Mô tả ngắn, Precondition, Input, Return, Exception, LogMsg, Type, P/F, Date)
    ("UTCID01", "Login thành công, trả về token",
     "User tồn tại, email xác nhận, tài khoản active",
     "email=admin1@homedecorshop.local\npassword=admin123",
     "200 + AuthResult(token, user)",
     "", "Đăng nhập thành công", "N", "P", "Jun-25"),

    ("UTCID02", "Sai mật khẩu → null",
     "User tồn tại, tài khoản active",
     "email=admin1@homedecorshop.local\npassword=WrongPassword_123",
     "null", "", "Email hoặc mật khẩu không chính xác", "A", "P", "Jun-25"),

    ("UTCID03", "Email không tồn tại → null",
     "Email chưa đăng ký trong hệ thống",
     "email=notexist@homedecorshop.local\npassword=Password123",
     "null", "", "Email hoặc mật khẩu không chính xác", "A", "P", "Jun-25"),

    ("UTCID04", "Email chưa xác nhận → Exception",
     "User tồn tại, IsEmailConfirmed=false",
     "email=unconfirmed@homedecorshop.local\npassword=Password123",
     "", "RequestValidationException (EmailNotConfirmed)",
     "Email chưa được xác nhận", "A", "P", "Jun-25"),

    ("UTCID05", "Tài khoản bị khóa → Exception",
     "User tồn tại, IsActive=false",
     "email=locked@homedecorshop.local\npassword=Password123",
     "", "RequestValidationException",
     "Tài khoản của bạn đã bị khóa", "A", "P", "Jun-25"),

    ("UTCID06", "Login thành công → Update token trong DB",
     "User tồn tại, đủ điều kiện đăng nhập",
     "email=admin1@homedecorshop.local\npassword=admin123",
     "Repo.Update gọi đúng 1 lần", "", "Token được cập nhật vào DB", "N", "P", "Jun-25"),

    ("UTCID07", "Login thành công → sinh token mới khác token cũ",
     "User tồn tại, token cũ đang rỗng",
     "email=admin1@homedecorshop.local\npassword=admin123",
     "result.Token != null && != oldToken", "",
     "Token mới được tạo khác token cũ", "B", "P", "Jun-25"),

    ("UTCID08", "Email không phân biệt hoa/thường + có khoảng trắng",
     "User tồn tại với email lowercase",
     "email=  Admin1@HomeDecorShop.LOCAL  \npassword=admin123",
     "200 + AuthResult đúng user", "",
     "Hệ thống chuẩn hóa email trước khi xử lý", "B", "P", "Jun-25"),
]
SHEETS.append(("1.Login", "Đăng nhập (Login)", "Tạ Đức Bảo", 150, "HOM-3", login_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 2. REGISTER (UserServiceRegisterTests)
# ══════════════════════════════════════════════════════════════════════════════
reg_tcs = [
    ("UTCID01", "Đăng ký thành công, trả về token",
     "Email chưa tồn tại trong hệ thống",
     "email=newuser@homedecorshop.local\nfullName=Nguyen Van A\nphone=0987654321\npassword=Password123\nrole=customer",
     "200 + AuthResult(token, user)", "",
     "Tài khoản được tạo thành công", "N", "P", "Jun-25"),

    ("UTCID02", "Email đã tồn tại → ConflictException",
     "Email admin1@homedecorshop.local đã được đăng ký",
     "email=admin1@homedecorshop.local\nfullName=Test\npassword=Password123\nrole=customer",
     "", "ConflictException (EmailAlreadyExists)",
     "Email đã được sử dụng", "A", "P", "Jun-25"),

    ("UTCID03", "Email được chuẩn hóa về lowercase + trim",
     "Email chưa tồn tại",
     "email=  UserMixed@Example.COM  \nfullName=Test\npassword=Password123\nrole=customer",
     "result.User.Email = 'usermixed@example.com'", "",
     "Email được chuẩn hóa trước khi lưu", "B", "P", "Jun-25"),

    ("UTCID04", "Mật khẩu được hash, không lưu plaintext",
     "Email chưa tồn tại",
     "email=hash_test@example.com\npassword=Password123\nrole=customer",
     "capturedUser.PasswordHash != 'Password123'\nBCrypt.Verify = true", "",
     "Mật khẩu được hash bằng BCrypt", "N", "P", "Jun-25"),

    ("UTCID05", "User mới mặc định có role Customer",
     "Email chưa tồn tại",
     "email=role_test@example.com\nrole=customer",
     "result.User.Role = 'customer'", "",
     "Role mặc định được gán là customer", "N", "P", "Jun-25"),

    ("UTCID06", "Repository.Create được gọi đúng 1 lần",
     "Email chưa tồn tại",
     "email=verify_call@example.com\npassword=Password123",
     "Repo.Create gọi Times.Once", "",
     "Repo.Create được gọi chính xác 1 lần", "N", "P", "Jun-25"),

    ("UTCID07", "FullName được trim khoảng trắng",
     "Email chưa tồn tại",
     "email=trim_test@example.com\nfullName=  Nguyen Van B  \npassword=Password123",
     "result.User.FullName = 'Nguyen Van B'", "",
     "FullName được trim trước khi lưu", "B", "P", "Jun-25"),
]
SHEETS.append(("2.Register", "Đăng ký tài khoản (Register)", "Tạ Đức Bảo", 120, "HOM-2", reg_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 3. ADDRESS (UserServiceAddressTests)
# ══════════════════════════════════════════════════════════════════════════════
addr_tcs = [
    ("UTCID01", "GetAddresses - Token hợp lệ, trả về danh sách địa chỉ",
     "User có 2 địa chỉ trong DB",
     "token=valid-token-123",
     "200 + List<Address> (count=2)", "", "Trả về 2 địa chỉ", "N", "P", "Jun-25"),

    ("UTCID02", "GetAddresses - Token không hợp lệ → null",
     "Token không khớp bất kỳ user nào",
     "token=invalid-token",
     "null", "", "Không tìm thấy user", "A", "P", "Jun-25"),

    ("UTCID03", "GetAddresses - Địa chỉ mặc định đứng đầu danh sách",
     "User có 2 địa chỉ, 1 IsDefault=true",
     "token=valid-token-123",
     "result.First().IsDefault = true", "",
     "Địa chỉ mặc định được sắp xếp lên đầu", "N", "P", "Jun-25"),

    ("UTCID04", "GetAddressById - Địa chỉ tồn tại → trả về địa chỉ đúng",
     "User có address ID=1",
     "token=valid-token-123\naddressId=1",
     "200 + AddressView(Id=1, Line1='123 Le Loi')", "",
     "Trả về địa chỉ chính xác", "N", "P", "Jun-25"),

    ("UTCID05", "GetAddressById - ID không tồn tại → null",
     "User không có address ID=999",
     "token=valid-token-123\naddressId=999",
     "null", "", "Không tìm thấy địa chỉ", "A", "P", "Jun-25"),

    ("UTCID06", "GetAddressById - Token không hợp lệ → null",
     "Token không hợp lệ",
     "token=bad-token\naddressId=1",
     "null", "", "Token không hợp lệ", "A", "P", "Jun-25"),

    ("UTCID07", "AddAddress - Thêm địa chỉ mới thành công",
     "Token hợp lệ, user tồn tại",
     "token=valid\nfullName=Tran Van B\nline1=789 Tran Hung Dao\nIsDefault=false",
     "200 + AddressView(FullName='Tran Van B')", "",
     "Địa chỉ mới được thêm thành công", "N", "P", "Jun-25"),

    ("UTCID08", "AddAddress - Đặt làm mặc định → reset các địa chỉ khác",
     "User đã có 1 địa chỉ mặc định (ID=1)",
     "token=valid\nIsDefault=true",
     "address ID=1 có IsDefault=false", "",
     "Địa chỉ cũ được bỏ mặc định", "A", "P", "Jun-25"),

    ("UTCID09", "AddAddress - Token không hợp lệ → null",
     "Token không tìm thấy user",
     "token=invalid",
     "null", "", "Token không hợp lệ", "A", "P", "Jun-25"),

    ("UTCID10", "UpdateAddress - Cập nhật thành công",
     "User có address ID=1",
     "token=valid\naddressId=1\nfullName=Updated Name",
     "200 + AddressView(FullName='Updated Name')", "",
     "Cập nhật địa chỉ thành công", "N", "P", "Jun-25"),

    ("UTCID11", "UpdateAddress - ID không tồn tại → null",
     "User không có address ID=999",
     "token=valid\naddressId=999",
     "null", "", "Không tìm thấy địa chỉ để cập nhật", "A", "P", "Jun-25"),

    ("UTCID12", "DeleteAddress - Xóa thành công → true",
     "User có address ID=2",
     "token=valid\naddressId=2",
     "true", "", "Địa chỉ được xóa thành công", "N", "P", "Jun-25"),

    ("UTCID13", "DeleteAddress - ID không tồn tại → false",
     "User không có address ID=999",
     "token=valid\naddressId=999",
     "false", "", "Không tìm thấy địa chỉ để xóa", "A", "P", "Jun-25"),

    ("UTCID14", "DeleteAddress - Token không hợp lệ → false",
     "Token không hợp lệ",
     "token=bad-token\naddressId=1",
     "false", "", "Token không hợp lệ", "A", "P", "Jun-25"),
]
SHEETS.append(("3.Address", "Quản lý địa chỉ (Address)", "Tạ Đức Bảo", 200, "HOM-4", addr_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 4. CART (CartServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
cart_tcs = [
    ("UTCID01", "AddItem - Sản phẩm không hoạt động (IsActive=false) → Exception",
     "Sản phẩm ProductId=10 có IsActive=false",
     "token=valid\nproductId=10\nquantity=2",
     "", "ConflictException (ProductInactive)",
     "Sản phẩm không còn hoạt động", "A", "P", "Jun-25"),

    ("UTCID02", "AddItem - Sản phẩm hết hàng (InStock=false) → Exception",
     "Sản phẩm ProductId=11 có InStock=false, StockLeft=0",
     "token=valid\nproductId=11\nquantity=1",
     "", "ConflictException (ProductOutOfStock)",
     "Sản phẩm đã hết hàng", "A", "P", "Jun-25"),

    ("UTCID03", "AddItem - Số lượng vượt tồn kho → Exception",
     "Sản phẩm StockLeft=3",
     "token=valid\nproductId=12\nquantity=5",
     "", "ConflictException (ProductStockExceeded)",
     "Số lượng yêu cầu vượt quá tồn kho", "B", "P", "Jun-25"),

    ("UTCID04", "UpdateItem - Cập nhật số lượng vượt tồn kho → Exception",
     "CartItem tồn tại, StockLeft=3",
     "token=valid\ncartItemId=100\nquantity=5",
     "", "ConflictException (ProductStockExceeded)",
     "Số lượng cập nhật vượt quá tồn kho", "B", "P", "Jun-25"),

    ("UTCID05", "UpdateItem - Cập nhật số lượng hợp lệ → thành công",
     "CartItem tồn tại, StockLeft=5",
     "token=valid\ncartItemId=100\nquantity=3",
     "200 + CartView (items[0].Quantity=3)", "",
     "Số lượng giỏ hàng được cập nhật", "N", "P", "Jun-25"),

    ("UTCID06", "RemoveItem - Xóa sản phẩm khỏi giỏ hàng thành công",
     "CartItem ID=100 tồn tại trong giỏ hàng",
     "token=valid\ncartItemId=100",
     "true + cart.Items rỗng", "",
     "Sản phẩm đã được xóa khỏi giỏ hàng", "N", "P", "Jun-25"),

    ("UTCID07", "Clear - Làm trống giỏ hàng thành công",
     "Giỏ hàng có 2 sản phẩm",
     "token=valid",
     "true + cart.Items.Count=0", "",
     "Giỏ hàng đã được làm trống", "N", "P", "Jun-25"),
]
SHEETS.append(("4.Cart", "Giỏ hàng (Cart)", "Tạ Đức Bảo", 180, "HOM-6", cart_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 5. ORDER (OrderServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
order_tcs = [
    ("UTCID01", "PlaceOrder - Đặt hàng thành công, tính đúng tiền, giảm stock, xóa cart",
     "Giỏ hàng có 1 sản phẩm (qty=2, price=100k), StockLeft=5",
     "token=valid\nfullName=Nguyen Van A\nphone=0123456789\ncity=HCM",
     "Subtotal=200k, ShippingFee=30k, Total=230k\nStockLeft giảm còn 3\ncart.Items rỗng",
     "", "Đặt hàng thành công", "N", "P", "Jun-25"),

    ("UTCID02", "PlaceOrder - Giỏ hàng rỗng → Exception",
     "Giỏ hàng user không có sản phẩm",
     "token=valid",
     "", "RequestValidationException",
     "Giỏ hàng không được để trống", "A", "P", "Jun-25"),

    ("UTCID03", "PlaceOrder - Sản phẩm không đủ tồn kho → Exception",
     "Sản phẩm StockLeft=1, đặt mua qty=2",
     "token=valid\nproductId=10\nquantity=2",
     "", "ConflictException",
     "Không đủ hàng trong kho để xử lý đơn", "B", "P", "Jun-25"),

    ("UTCID04", "Cancel - Hủy đơn hàng thành công, hoàn kho, cập nhật trạng thái",
     "Đơn hàng Status=PendingPayment, không có giao dịch VNPay pending",
     "token=valid\norderId=100",
     "result.Status='cancelled'\nStockLeft tăng từ 10 lên 13",
     "", "Đơn hàng đã được hủy", "N", "P", "Jun-25"),

    ("UTCID05", "Cancel - Đơn có giao dịch VNPay đang Pending → Exception",
     "Đơn hàng có Payment với method=vnpay, status=Pending",
     "token=valid\norderId=100",
     "", "ConflictException",
     "Không thể hủy đơn khi đang có giao dịch VNPay chờ xử lý", "A", "P", "Jun-25"),

    ("UTCID06", "RequestRefund - Đơn đã thanh toán → chuyển trạng thái RefundRequested",
     "Đơn hàng PaymentStatus=Paid, Status=Completed",
     "token=valid\norderId=200\nreason=Sản phẩm lỗi",
     "result.Status='refund_requested'\norder.Notes chứa '[KHIẾU NẠI]: Sản phẩm lỗi'",
     "", "Yêu cầu hoàn hàng đã được gửi", "N", "P", "Jun-25"),

    ("UTCID07", "RequestRefund - Đơn chưa thanh toán → Exception",
     "Đơn hàng PaymentStatus=Pending",
     "token=valid\norderId=200",
     "", "ConflictException",
     "Chỉ có thể yêu cầu hoàn tiền với đơn đã thanh toán", "A", "P", "Jun-25"),

    ("UTCID08", "ProcessRefund - Admin duyệt hoàn tiền → gọi WalletService, status=Refunded",
     "Admin token, đơn Status=RefundRequested, Total=500k",
     "token=admin\norderId=300\napprove=true",
     "result.Status='refunded'\nPaymentStatus=Refunded\nWalletService.ProcessRefundPayment gọi 1 lần",
     "", "Hoàn tiền đã được phê duyệt", "N", "P", "Jun-25"),

    ("UTCID09", "ProcessRefund - Admin từ chối → status trở về Completed",
     "Admin token, đơn Status=RefundRequested",
     "token=admin\norderId=300\napprove=false",
     "result.Status='completed'\nWalletService KHÔNG được gọi",
     "", "Yêu cầu hoàn hàng đã bị từ chối", "N", "P", "Jun-25"),
]
SHEETS.append(("5.Order", "Đặt hàng & Hoàn hàng (Order)", "Tạ Đức Bảo", 250, "HOM-7", order_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 6. PRODUCT (ProductServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
product_tcs = [
    ("UTCID01", "Search - Từ khóa có khoảng trắng/hoa thường → chuẩn hóa + tìm đúng",
     "DB có 3 sản phẩm (Chair, Desk, Sofa)",
     "query='  chair  '\npage=1\npageSize=10",
     "Total=1, Items chứa Product 'Classic Wooden Chair'", "",
     "Tìm kiếm chuẩn hóa từ khóa thành công", "B", "P", "Jun-25"),

    ("UTCID02", "Search - Lọc theo khoảng giá MinPrice/MaxPrice",
     "DB có 3 sản phẩm (giá 120, 250, 450)",
     "minPrice=100\nmaxPrice=300",
     "Total=2, Products ID=1 (120) và ID=2 (250)", "",
     "Lọc theo khoảng giá hoạt động chính xác", "N", "P", "Jun-25"),

    ("UTCID03", "Search - Lọc chỉ sản phẩm còn hàng (InStockOnly=true)",
     "Product 3 (Sofa) có InStock=false",
     "inStockOnly=true",
     "Total=2, không chứa Product ID=3", "",
     "Lọc sản phẩm còn hàng hoạt động chính xác", "N", "P", "Jun-25"),

    ("UTCID04", "Search - Lọc sản phẩm đang giảm giá (OnSaleOnly=true)",
     "Product 2 (Desk) không có OldPrice",
     "onSaleOnly=true",
     "Total=2, chứa Product ID=1,3, không chứa ID=2", "",
     "Lọc sản phẩm giảm giá hoạt động chính xác", "N", "P", "Jun-25"),

    ("UTCID05", "Search - Sắp xếp theo giá tăng dần (price-asc)",
     "DB có 3 sản phẩm (giá 120, 250, 450)",
     "sortBy=price-asc",
     "Thứ tự: ID=1(120) → ID=2(250) → ID=3(450)", "",
     "Sắp xếp giá tăng dần chính xác", "N", "P", "Jun-25"),

    ("UTCID06", "Search - Sắp xếp theo giá giảm dần (price-desc)",
     "DB có 3 sản phẩm",
     "sortBy=price-desc",
     "Thứ tự: ID=3(450) → ID=2(250) → ID=1(120)", "",
     "Sắp xếp giá giảm dần chính xác", "N", "P", "Jun-25"),

    ("UTCID07", "Search - Sắp xếp theo rating giảm dần (rating-desc)",
     "Rating: ID=2(4.8), ID=1(4.5), ID=3(4.0)",
     "sortBy=rating-desc",
     "Thứ tự: ID=2 → ID=1 → ID=3", "",
     "Sắp xếp theo rating chính xác", "N", "P", "Jun-25"),

    ("UTCID08", "Search - Sắp xếp mới nhất (newest)",
     "CreatedAt: ID=2(-2 ngày), ID=3(-5 ngày), ID=1(-10 ngày)",
     "sortBy=newest",
     "Thứ tự: ID=2 → ID=3 → ID=1", "",
     "Sắp xếp mới nhất chính xác", "N", "P", "Jun-25"),

    ("UTCID09", "Search - Sắp xếp theo độ liên quan (relevance)",
     "Query='Modern', ID=2 có rating 4.8 > ID=3 rating 4.0",
     "query=Modern\nsortBy=relevance",
     "Total=2, ID=2 đứng trước ID=3", "",
     "Xếp hạng độ liên quan chính xác", "N", "P", "Jun-25"),

    ("UTCID10", "Search - Phân trang đúng (Page=2, PageSize=2)",
     "DB có 3 sản phẩm",
     "sortBy=price-asc\npage=2\npageSize=2",
     "Total=3, Items.Count=1, Items[0].ID=3", "",
     "Phân trang hoạt động chính xác", "N", "P", "Jun-25"),

    ("UTCID11", "GetReviews - Trả về danh sách reviews đúng mapping",
     "ProductId=1 có 2 reviews trong DB",
     "productId=1",
     "result.Count=2\nfirst.Author='Nguyen Van A', first.Rating=5", "",
     "GetReviews trả về đúng dữ liệu", "N", "P", "Jun-25"),

    ("UTCID12", "AddReview - Sản phẩm không tồn tại → Exception",
     "ProductId=999 không có trong DB",
     "productId=999\nauthor=Khách\nrating=5",
     "", "Exception('Product does not exist.')",
     "Không thêm review khi sản phẩm không tồn tại", "A", "P", "Jun-25"),

    ("UTCID13", "AddReview - Rating > 5 bị clamp về 5",
     "ProductId=1 tồn tại",
     "productId=1\nauthor=User C\nrating=6",
     "result.Rating=5\nproduct.Reviews=3\nproduct.Rating cập nhật chính xác", "",
     "Rating được giới hạn tối đa bằng 5", "B", "P", "Jun-25"),

    ("UTCID14", "AddReview - Rating < 1 bị clamp về 1",
     "ProductId=1 tồn tại",
     "productId=1\nauthor=User C\nrating=0",
     "result.Rating=1", "",
     "Rating được giới hạn tối thiểu bằng 1", "B", "P", "Jun-25"),

    ("UTCID15", "AddReview - Chưa có review nào → Rating và Count được khởi tạo đúng",
     "ProductId=1, chưa có review nào trong DB",
     "productId=1\nauthor=User D\nrating=5",
     "product.Rating=5.0\nproduct.Reviews=1", "",
     "Product được cập nhật rating khi thêm review đầu tiên", "N", "P", "Jun-25"),
]
SHEETS.append(("6.Product", "Sản phẩm & Đánh giá (Product)", "Tạ Đức Bảo", 350, "HOM-8", product_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 7. CATEGORY (CategoryServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
cat_tcs = [
    ("UTCID01", "Update - Không cho tắt danh mục có sản phẩm đang bán → Exception",
     "Category ID=1 HasActiveProducts=true",
     "categoryId=1\nisActive=false",
     "", "ConflictException (CategoryHasActiveProducts)",
     "Không thể tắt danh mục đang có sản phẩm hoạt động", "A", "P", "Jun-25"),

    ("UTCID02", "Delete - Không xóa danh mục chứa sản phẩm → HasProducts",
     "Category ID=1 HasProducts=true",
     "categoryId=1",
     "CategoryDeleteResult.HasProducts", "",
     "Không xóa được danh mục đang có sản phẩm", "A", "P", "Jun-25"),

    ("UTCID03", "Create - Tên danh mục trùng (case-insensitive + trim) → Exception",
     "Đã có category tên 'Chairs' (ID=1)",
     "name='  chairs  '\nslug=new-slug\ngroupId=1",
     "", "ConflictException (CategoryNameAlreadyExists)",
     "Tên danh mục đã được sử dụng", "A", "P", "Jun-25"),

    ("UTCID04", "Create - Slug trùng (case-insensitive + trim) → Exception",
     "Đã có category slug 'chairs' (ID=1)",
     "name=New Name\nslug='  ChAiRs  '\ngroupId=1",
     "", "ConflictException (CategorySlugAlreadyExists)",
     "Slug danh mục đã được sử dụng", "A", "P", "Jun-25"),

    ("UTCID05", "Update - Tên trùng với category khác → Exception",
     "Category ID=2 (Tables) đổi tên sang 'Chairs' đã tồn tại",
     "categoryId=2\nname=Chairs\nslug=tables-new",
     "", "ConflictException (CategoryNameAlreadyExists)",
     "Không thể đặt tên trùng với danh mục khác", "A", "P", "Jun-25"),

    ("UTCID06", "Update - Slug trùng với category khác → Exception",
     "Category ID=2 đổi slug sang 'chairs' đã tồn tại",
     "categoryId=2\nname=Tables New\nslug=chairs",
     "", "ConflictException (CategorySlugAlreadyExists)",
     "Không thể đặt slug trùng với danh mục khác", "A", "P", "Jun-25"),
]
SHEETS.append(("7.Category", "Danh mục sản phẩm (Category)", "Tạ Đức Bảo", 140, "HOM-9", cat_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 8. WALLET (WalletServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
wallet_tcs = [
    ("UTCID01", "Deposit - Nạp tiền hợp lệ, số dư tăng đúng",
     "User có ví số dư 1000, token hợp lệ",
     "token=valid\namount=500",
     "result.Balance=1500\nRepo.Update gọi 1 lần với Balance=1500\nTransaction Type=Deposit, Status=Success",
     "", "Nạp tiền thành công", "N", "P", "Jun-25"),

    ("UTCID02", "Deposit - Amount <= 0 → RequestValidationException",
     "Token hợp lệ",
     "token=valid\namount=0",
     "", "RequestValidationException (amount)",
     "Số tiền nạp phải lớn hơn 0", "B", "P", "Jun-25"),

    ("UTCID03", "Deposit - Amount âm → RequestValidationException",
     "Token hợp lệ",
     "token=valid\namount=-100",
     "", "RequestValidationException (amount)",
     "Số tiền không được âm", "B", "P", "Jun-25"),

    ("UTCID04", "Deposit - Ví chưa tồn tại → tạo ví mới rồi nạp",
     "User chưa có ví",
     "token=valid\namount=1000",
     "result.Balance=1000\nRepo.Create gọi 1 lần\nTransaction được tạo", "",
     "Ví mới được tạo và nạp tiền thành công", "N", "P", "Jun-25"),

    ("UTCID05", "Deposit - Token không hợp lệ → UnauthorizedException",
     "Token không khớp bất kỳ user nào",
     "token=wrong-token\namount=1000",
     "", "UnauthorizedException (AuthTokenInvalid)",
     "Xác thực token thất bại", "A", "P", "Jun-25"),

    ("UTCID06", "Withdraw - Rút tiền hợp lệ, số dư giảm đúng",
     "User có ví số dư 2000",
     "token=valid\namount=700",
     "result.Balance=1300\nTransaction Type=Withdraw, Status=Success", "",
     "Rút tiền thành công", "N", "P", "Jun-25"),

    ("UTCID07", "Withdraw - Amount <= 0 → RequestValidationException",
     "Token hợp lệ",
     "token=valid\namount=0",
     "", "RequestValidationException (amount)",
     "Số tiền rút phải lớn hơn 0", "B", "P", "Jun-25"),

    ("UTCID08", "Withdraw - Số tiền rút vượt số dư → ConflictException",
     "User có ví số dư 500",
     "token=valid\namount=1000",
     "", "ConflictException (WalletInsufficientBalance)",
     "Số dư không đủ để thực hiện rút tiền", "A", "P", "Jun-25"),

    ("UTCID09", "Withdraw - Token không hợp lệ → UnauthorizedException",
     "Token sai",
     "token=wrong-token\namount=100",
     "", "UnauthorizedException (AuthTokenInvalid)",
     "Xác thực token thất bại", "A", "P", "Jun-25"),

    ("UTCID10", "PayOrder - Thanh toán đơn hàng hợp lệ bằng ví",
     "Ví user có 500k, đơn hàng 400k, Status=PendingPayment",
     "token=valid\norderId=1",
     "result.Balance=100k\nPayment tạo với method=wallet, status=Paid\nOrder chuyển sang Processing\nAdmin ví tăng 400k",
     "", "Thanh toán bằng ví thành công", "N", "P", "Jun-25"),

    ("UTCID11", "PayOrder - Đơn hàng không tồn tại → NotFoundException",
     "",
     "token=valid\norderId=999",
     "", "NotFoundException (OrderNotFound)",
     "Không tìm thấy đơn hàng", "A", "P", "Jun-25"),

    ("UTCID12", "PayOrder - Đơn hàng của người khác → NotFoundException",
     "Đơn hàng thuộc UserId=999",
     "token=valid (userId=10)\norderId=2",
     "", "NotFoundException (OrderNotFound)",
     "Không có quyền thanh toán đơn hàng của người khác", "A", "P", "Jun-25"),

    ("UTCID13", "PayOrder - Đơn hàng đã hủy → ConflictException",
     "Đơn hàng Status=Cancelled",
     "token=valid\norderId=3",
     "", "ConflictException (OrderCancelled)",
     "Không thể thanh toán đơn đã hủy", "A", "P", "Jun-25"),

    ("UTCID14", "PayOrder - Đơn hàng đã thanh toán → ConflictException",
     "Đơn hàng PaymentStatus=Paid",
     "token=valid\norderId=4",
     "", "ConflictException (OrderAlreadyPaid)",
     "Đơn hàng đã được thanh toán trước đó", "A", "P", "Jun-25"),

    ("UTCID15", "PayOrder - Số dư ví không đủ → ConflictException",
     "Ví có 100k, đơn hàng 500k",
     "token=valid\norderId=6",
     "", "ConflictException (WalletInsufficientBalance)",
     "Số dư ví không đủ để thanh toán đơn hàng", "A", "P", "Jun-25"),

    ("UTCID16", "PayOrder - Token không hợp lệ → UnauthorizedException",
     "",
     "token=wrong-token\norderId=1",
     "", "UnauthorizedException (AuthTokenInvalid)",
     "Xác thực token thất bại", "A", "P", "Jun-25"),
]
SHEETS.append(("8.Wallet", "Ví điện tử (Wallet)", "Tạ Đức Bảo", 400, "HOM-10", wallet_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 9. PAYMENT (PaymentServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
pay_tcs = [
    ("UTCID01", "CreateVnPayPayment - Đơn hàng hợp lệ → tạo Payment Pending VNPay",
     "Đơn hàng ID=1 hợp lệ, Status=PendingPayment, chưa có payment",
     "token=valid\norderId=1",
     "result.PaymentId=1\nresult.TransactionCode bắt đầu bằng 'VNPAY1'\nPayment tạo với method=vnpay, status=Pending",
     "", "Tạo giao dịch VNPay thành công", "N", "P", "Jun-25"),

    ("UTCID02", "CreateVnPayPayment - Đã có Payment VNPay Pending → dùng lại",
     "Đơn hàng đã có payment VNPay pending với code an toàn",
     "token=valid\norderId=2",
     "result.PaymentId=5 (payment cũ)\nrepo.Create KHÔNG được gọi", "",
     "Tái sử dụng payment VNPay đang chờ xử lý", "N", "P", "Jun-25"),

    ("UTCID03", "CreateVnPayPayment - TransactionCode không an toàn → tạo lại code mới",
     "Payment cũ có code chứa ký tự đặc biệt 'VNPAY-UNSAFE-CODE'",
     "token=valid\norderId=3",
     "result.TransactionCode bắt đầu bằng 'VNPAY3', chỉ chứa chữ/số\nrepo.Update gọi 1 lần",
     "", "TransactionCode không an toàn được tái tạo", "A", "P", "Jun-25"),

    ("UTCID04", "CreateVnPayPayment - Đơn hàng không tồn tại → NotFoundException",
     "",
     "token=valid\norderId=999",
     "", "NotFoundException (OrderNotFound)",
     "Không tìm thấy đơn hàng", "A", "P", "Jun-25"),

    ("UTCID05", "CreateVnPayPayment - Đơn hàng của người khác → NotFoundException",
     "Đơn hàng thuộc user khác",
     "token=valid\norderId=4",
     "", "NotFoundException (OrderNotFound)",
     "Không có quyền tạo thanh toán cho đơn của người khác", "A", "P", "Jun-25"),

    ("UTCID06", "CreateVnPayPayment - Đơn hàng đã hủy → ConflictException",
     "Đơn hàng Status=Cancelled",
     "token=valid\norderId=5",
     "", "ConflictException (OrderCancelled)",
     "Không thể thanh toán đơn đã hủy", "A", "P", "Jun-25"),

    ("UTCID07", "CreateVnPayPayment - Đơn hàng đã thanh toán → ConflictException",
     "Đơn hàng PaymentStatus=Paid",
     "token=valid\norderId=6",
     "", "ConflictException (OrderAlreadyPaid)",
     "Đơn hàng đã được thanh toán", "A", "P", "Jun-25"),

    ("UTCID08", "CreateVnPayPayment - Đơn hàng không ở trạng thái PendingPayment → ConflictException",
     "Đơn hàng Status=Processing",
     "token=valid\norderId=7",
     "", "ConflictException (OrderPaymentNotPending)",
     "Đơn hàng không ở trạng thái chờ thanh toán", "A", "P", "Jun-25"),

    ("UTCID09", "CreateVnPayPayment - Token không hợp lệ → UnauthorizedException",
     "",
     "token=wrong-token\norderId=1",
     "", "UnauthorizedException (AuthTokenInvalid)",
     "Xác thực token thất bại", "A", "P", "Jun-25"),

    ("UTCID10", "HandleVnPayCallback - Thanh toán thành công → cập nhật Payment+Order, nạp ví Admin",
     "Payment tồn tại, ResponseCode=00, TransactionStatus=00, số tiền khớp",
     "transactionCode=VNPAY8SUCCESS\namount=30000000\nresponseCode=00",
     "result.IsSuccess=true\nresult.PaymentStatus='paid'\nresult.OrderStatus='processing'\nWalletService.AddToAdminWallet gọi 1 lần",
     "", "Thanh toán VNPay thành công được xử lý", "N", "P", "Jun-25"),

    ("UTCID11", "HandleVnPayCallback - Thanh toán thất bại (ResponseCode≠00) → PaymentFailed",
     "Payment tồn tại, ResponseCode=24",
     "transactionCode=VNPAY9FAILED\nresponseCode=24",
     "result.IsSuccess=false\nresult.PaymentStatus='failed'\nOrder KHÔNG được cập nhật\nWalletService KHÔNG được gọi",
     "", "Thanh toán VNPay thất bại được xử lý đúng", "A", "P", "Jun-25"),

    ("UTCID12", "HandleVnPayCallback - Số tiền không khớp → ConflictException",
     "Payment Amount=300k nhưng callback gửi về 100k",
     "transactionCode=VNPAY10AMOUNT\namount=10000000\nresponseCode=00",
     "", "ConflictException (PaymentGatewayAmountInvalid)",
     "Số tiền callback không khớp với đơn hàng", "A", "P", "Jun-25"),

    ("UTCID13", "HandleVnPayCallback - TransactionCode không tồn tại → NotFoundException",
     "",
     "transactionCode=UNKNOWN\nresponseCode=00",
     "", "NotFoundException (PaymentNotFound)",
     "Không tìm thấy giao dịch thanh toán", "A", "P", "Jun-25"),

    ("UTCID14", "HandleVnPayCallback - Payment không phải VNPay → ConflictException",
     "Payment method=wallet",
     "transactionCode=WALLET123\nresponseCode=00",
     "", "ConflictException (PaymentGatewayCallbackInvalid)",
     "Giao dịch không phải loại VNPay", "A", "P", "Jun-25"),

    ("UTCID15", "HandleVnPayCallback - Callback trùng lặp (đã Paid) → bỏ qua, trả thành công",
     "Payment đã có Status=Paid",
     "transactionCode=VNPAY12DUPLICATE\nresponseCode=00",
     "result.IsSuccess=true\nrepo.Update KHÔNG được gọi\nWalletService KHÔNG được gọi",
     "", "Callback trùng lặp được bỏ qua an toàn", "B", "P", "Jun-25"),
]
SHEETS.append(("9.Payment", "Thanh toán VNPay (Payment)", "Tạ Đức Bảo", 450, "HOM-11", pay_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 10. FEEDBACK (FeedbackServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
feedback_tcs = [
    ("UTCID01", "GetAll - Trả về danh sách feedback đúng mapping",
     "DB có 2 feedback",
     "",
     "result.Count=2\nresult.First().Name='User A'\nresult.Last().Email='b@gmail.com'",
     "", "Trả về đầy đủ danh sách feedback", "N", "P", "Jun-25"),

    ("UTCID02", "GetById - ID tồn tại → trả về FeedbackView đúng",
     "Feedback ID=10 tồn tại",
     "feedbackId=10",
     "result.FeedbackId=10\nresult.Name='Nguyen Van A'",
     "", "Trả về feedback đúng ID", "N", "P", "Jun-25"),

    ("UTCID03", "GetById - ID không tồn tại → null",
     "",
     "feedbackId=999",
     "null", "", "Không tìm thấy feedback", "A", "P", "Jun-25"),

    ("UTCID04", "Create - Trim khoảng trắng và lowercase email khi tạo",
     "",
     "name='   Trần Văn B   '\nemail='  bAnG@HomeDecor.COM  '\nmessage='  Sản phẩm rất đẹp!  '",
     "result.Name='Trần Văn B'\nresult.Email='bang@homedecor.com'\nresult.Message='Sản phẩm rất đẹp!'",
     "", "Dữ liệu được chuẩn hóa trước khi lưu", "N", "P", "Jun-25"),

    ("UTCID05", "Update - Giữ nguyên CreatedAt gốc khi cập nhật",
     "Feedback ID=5, CreatedAt=5 ngày trước",
     "feedbackId=5\nname=New Name\nemail=new@gmail.com\nmessage=New Message",
     "result.Name='New Name'\nrepo.Update gọi với CreatedAt=originalCreatedAt (không bị ghi đè)",
     "", "CreatedAt không bị thay đổi khi cập nhật", "N", "P", "Jun-25"),

    ("UTCID06", "Update - Feedback không tồn tại → null, không gọi repo.Update",
     "",
     "feedbackId=99",
     "null\nrepo.Update KHÔNG được gọi",
     "", "Không cập nhật khi không tìm thấy feedback", "A", "P", "Jun-25"),

    ("UTCID07", "Delete - Trả về kết quả từ repository (true/false)",
     "feedback ID=1 tồn tại, ID=2 không tồn tại",
     "feedbackId=1 → true\nfeedbackId=2 → false",
     "Delete(1)=true\nDelete(2)=false",
     "", "Xóa feedback theo kết quả từ repo", "N", "P", "Jun-25"),
]
SHEETS.append(("10.Feedback", "Phản hồi (Feedback)", "Tạ Đức Bảo", 120, "HOM-12", feedback_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 11. MARKETING – COUPON (MarketingServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
mkt_tcs = [
    ("UTCID01", "GetCouponsAsync - Có coupon → trả về danh sách CouponView",
     "DB có 2 coupon (SALE10, SALE20)",
     "",
     "result.Length=2\nresult[0].Code='SALE10', DiscountPercentage=10\nresult[1].Code='SALE20', DiscountPercentage=20",
     "", "Trả về danh sách coupon đầy đủ", "N", "P", "Jun-25"),

    ("UTCID02", "GetCouponsAsync - Không có coupon → trả về mảng rỗng",
     "DB không có coupon nào",
     "",
     "result không null + result.Length=0",
     "", "Trả về mảng rỗng khi không có coupon", "A", "P", "Jun-25"),

    ("UTCID03", "CreateCouponAsync - Tạo coupon hợp lệ, code uppercase và trim",
     "",
     "code=' sale10 '\ndiscountPercentage=10\nexpiryDate=+7 ngày\nmaxUsage=100",
     "result.Code='SALE10'\nresult.CurrentUsage=0\nresult.IsActive=true\nrepo.CreateCouponAsync gọi 1 lần",
     "", "Coupon được tạo thành công với code chuẩn hóa", "N", "P", "Jun-25"),

    ("UTCID04", "CreateCouponAsync - CreatedAt được đặt đúng thời điểm tạo",
     "",
     "code=NEWYEAR\ndiscountPercentage=15",
     "result.CreatedAt >= beforeCreate && result.CreatedAt <= afterCreate",
     "", "CreatedAt được gán chính xác khi tạo coupon", "N", "P", "Jun-25"),

    ("UTCID05", "ValidateCouponAsync - Coupon hợp lệ → trả về CouponView",
     "Coupon 'SALE10' đang active, chưa hết hạn, chưa hết lượt",
     "code='sale10' (lowercase)",
     "result không null\nresult.Code='SALE10'\nrepo.GetCouponByCodeAsync gọi với 'SALE10' (uppercase)",
     "", "Coupon hợp lệ được xác nhận thành công", "N", "P", "Jun-25"),

    ("UTCID06", "ValidateCouponAsync - Coupon không tồn tại → null",
     "",
     "code=notfound",
     "null", "", "Coupon không tồn tại", "A", "P", "Jun-25"),

    ("UTCID07", "ValidateCouponAsync - Coupon đã tắt (IsActive=false) → null",
     "Coupon 'INACTIVE' có IsActive=false",
     "code=inactive",
     "null", "", "Coupon đã bị tắt không hợp lệ", "A", "P", "Jun-25"),

    ("UTCID08", "ValidateCouponAsync - Coupon hết hạn (ExpiryDate < Now) → null",
     "Coupon 'EXPIRED' ExpiryDate=-1 ngày",
     "code=expired",
     "null", "", "Coupon đã hết hạn không hợp lệ", "A", "P", "Jun-25"),

    ("UTCID09", "ValidateCouponAsync - Coupon hết lượt dùng (CurrentUsage >= MaxUsage) → null",
     "Coupon 'LIMITED' CurrentUsage=MaxUsage=10",
     "code=limited",
     "null", "", "Coupon đã đạt giới hạn sử dụng", "B", "P", "Jun-25"),

    ("UTCID10", "DeleteCouponAsync - Coupon tồn tại → xóa thành công",
     "Coupon ID=1 tồn tại",
     "couponId=1",
     "repo.GetCouponByIdAsync(1) gọi 1 lần\nrepo.DeleteCouponAsync(coupon) gọi 1 lần",
     "", "Coupon được xóa thành công", "N", "P", "Jun-25"),

    ("UTCID11", "DeleteCouponAsync - Coupon không tồn tại → KeyNotFoundException",
     "",
     "couponId=999",
     "", "KeyNotFoundException('Coupon 999 not found')",
     "Không tìm thấy coupon để xóa", "A", "P", "Jun-25"),
]
SHEETS.append(("11.Marketing", "Mã giảm giá (Marketing/Coupon)", "Tạ Đức Bảo", 200, "HOM-13", mkt_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 12. DASHBOARD (DashboardServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
dash_tcs = [
    ("UTCID01", "GetStats - User không phải Admin → ForbiddenException",
     "User có Role=Customer",
     "token=dashboard_test_token (role=Customer)",
     "", "ForbiddenException('Bạn không có quyền truy cập thông tin này.')",
     "Chỉ Admin mới được xem thống kê", "A", "P", "Jun-25"),

    ("UTCID02", "GetStats - Token không hợp lệ (user=null) → ForbiddenException",
     "Token không khớp user nào",
     "token=invalid",
     "", "ForbiddenException",
     "Token không hợp lệ bị từ chối", "A", "P", "Jun-25"),

    ("UTCID03", "GetStats - Tính tăng trưởng doanh thu đúng (50%)",
     "Admin token. Tuần trước=10tr, Tuần này=15tr",
     "token=admin",
     "result.RevenueGrowthPercentage=50.0",
     "", "Tăng trưởng 50% được tính chính xác", "N", "P", "Jun-25"),

    ("UTCID04", "GetStats - Tuần trước = 0, tuần này > 0 → tăng trưởng 100%",
     "Admin token. Tuần trước không có doanh thu, tuần này có 5tr",
     "token=admin",
     "result.RevenueGrowthPercentage=100.0",
     "", "Tăng trưởng 100% khi tuần trước bằng 0", "B", "P", "Jun-25"),

    ("UTCID05", "GetStats - Cả 2 tuần doanh thu đều = 0 → 0%",
     "Admin token. Không có đơn hàng nào",
     "token=admin",
     "result.RevenueGrowthPercentage=0.0",
     "", "Tăng trưởng 0% khi không có doanh thu", "B", "P", "Jun-25"),

    ("UTCID06", "GetStats - Biểu đồ 7 ngày có đúng 7 điểm, khớp doanh thu theo ngày",
     "Admin token. Hôm nay: 2 đơn Paid (200k+300k). Hôm qua: 1 Paid(150k) + 1 Pending(999k)",
     "token=admin",
     "result.RevenueChart.Count=7\ntodayChartItem.Revenue=500000\nyesterdayChartItem.Revenue=150000",
     "", "Biểu đồ 7 ngày chính xác, chỉ tính đơn Paid", "N", "P", "Jun-25"),
]
SHEETS.append(("12.Dashboard", "Thống kê Dashboard (Admin)", "Tạ Đức Bảo", 150, "HOM-14", dash_tcs))

# ══════════════════════════════════════════════════════════════════════════════
# 13. SETTINGS (SettingsServiceTests)
# ══════════════════════════════════════════════════════════════════════════════
settings_tcs = [
    ("UTCID01", "GetSettingsAsync - Cài đặt tồn tại → trả về đúng",
     "DB có SystemSetting với Id=1",
     "",
     "result.StoreName='HomeDecorShop'\nresult.VatPercentage=10\nresult.DefaultShippingFee=30000",
     "", "Trả về cài đặt hệ thống đúng", "N", "P", "Jun-25"),

    ("UTCID02", "GetSettingsAsync - Cài đặt null → tạo cài đặt mặc định mới",
     "DB không có cài đặt",
     "",
     "result không null\nresult.Id=1\nrepo.UpdateSettingsAsync gọi 1 lần",
     "", "Cài đặt mặc định được tạo khi không tìm thấy", "A", "P", "Jun-25"),

    ("UTCID03", "UpdateSettingsAsync - Cập nhật thành công → trả về setting mới",
     "",
     "StoreName='HomeDecorShop Updated'\nVatPercentage=15\nDefaultShippingFee=50000",
     "result.StoreName='HomeDecorShop Updated'\nresult.VatPercentage=15\nresult.DefaultShippingFee=50000",
     "", "Cập nhật cài đặt thành công", "N", "P", "Jun-25"),

    ("UTCID04", "UpdateSettingsAsync - UpdatedAt được cập nhật đúng thời gian",
     "",
     "SystemSetting(Id=1, StoreName='Test Store')",
     "result.UpdatedAt trong khoảng [beforeUpdate, afterUpdate]",
     "", "Thời gian cập nhật được ghi nhận chính xác", "N", "P", "Jun-25"),

    ("UTCID05", "UpdateSettingsAsync - Repository được gọi đúng 1 lần",
     "",
     "SystemSetting(Id=1)",
     "repo.UpdateSettingsAsync gọi Times.Once",
     "", "Repository được gọi chính xác 1 lần", "N", "P", "Jun-25"),
]
SHEETS.append(("13.Settings", "Cài đặt hệ thống (Settings)", "Tạ Đức Bảo", 100, "HOM-15", settings_tcs))

# ─── BUILD EXCEL ──────────────────────────────────────────────────────────────
def build_sheet(wb, sheet_info):
    sheet_name, func_name, created_by, loc, jira, tcs = sheet_info

    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    # ---------- Column widths ----------
    col_widths = [4, 14, 22, 35, 35, 30, 35, 5, 5, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ======== ROW 1: Header info ========
    for c in range(1, 11):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=GREY_BG)
        ws.cell(row=1, column=c).border = thin_border()

    ws.merge_cells("A1:B1")
    ws.cell(row=1, column=1, value="Function Code").font = dark_font(bold=True)
    ws.cell(row=1, column=1).alignment = left_align()
    ws.cell(row=1, column=1).border = thin_border()

    ws.merge_cells("C1:D1")
    ws.cell(row=1, column=3, value="Function Name").font = dark_font(bold=True)
    ws.cell(row=1, column=3).alignment = left_align()
    ws.cell(row=1, column=3).border = thin_border()

    ws.merge_cells("E1:J1")
    ws.cell(row=1, column=5, value=func_name).font = dark_font(bold=False)
    ws.cell(row=1, column=5).alignment = left_align()
    ws.cell(row=1, column=5).border = thin_border()

    # ======== ROW 2 ========
    for c in range(1, 11):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=GREY_BG)
        ws.cell(row=2, column=c).border = thin_border()

    ws.merge_cells("A2:B2")
    ws.cell(row=2, column=1, value="Created By").font = dark_font(bold=True)
    ws.cell(row=2, column=1).alignment = left_align()
    ws.cell(row=2, column=1).border = thin_border()

    ws.cell(row=2, column=3, value=created_by).font = dark_font()
    ws.cell(row=2, column=3).alignment = left_align()
    ws.cell(row=2, column=3).border = thin_border()

    ws.merge_cells("D2:E2")
    ws.cell(row=2, column=4, value="Executed By").font = dark_font(bold=True)
    ws.cell(row=2, column=4).alignment = left_align()
    ws.cell(row=2, column=4).border = thin_border()

    ws.merge_cells("F2:J2")
    ws.cell(row=2, column=6).border = thin_border()

    # ======== ROW 3 ========
    for c in range(1, 11):
        ws.cell(row=3, column=c).fill = PatternFill("solid", fgColor=GREY_BG)
        ws.cell(row=3, column=c).border = thin_border()

    ws.merge_cells("A3:B3")
    ws.cell(row=3, column=1, value="Lines of code").font = dark_font(bold=True)
    ws.cell(row=3, column=1).alignment = left_align()
    ws.cell(row=3, column=1).border = thin_border()

    ws.cell(row=3, column=3, value=loc).font = dark_font()
    ws.cell(row=3, column=3).alignment = center_align()
    ws.cell(row=3, column=3).border = thin_border()

    ws.merge_cells("D3:E3")
    ws.cell(row=3, column=4, value="Jira Ticket").font = dark_font(bold=True)
    ws.cell(row=3, column=4).alignment = left_align()
    ws.cell(row=3, column=4).border = thin_border()

    ws.merge_cells("F3:J3")
    ws.cell(row=3, column=6, value=jira).font = dark_font()
    ws.cell(row=3, column=6).alignment = left_align()
    ws.cell(row=3, column=6).border = thin_border()

    # ======== ROW 4: Test requirement ========
    ws.merge_cells("A4:J4")
    cell4 = ws.cell(row=4, column=1, value="Test requirement")
    cell4.fill = hdr_fill(DARK_BLUE)
    cell4.font = white_font(bold=True, size=10)
    cell4.alignment = left_align()
    cell4.border = thin_border()

    # ======== ROW 5: Summary labels ========
    n_tcs = len(tcs)
    n_pass = sum(1 for t in tcs if t[8] == "P")
    n_fail = n_tcs - n_pass
    type_n = sum(1 for t in tcs if t[7] == "N")
    type_a = sum(1 for t in tcs if t[7] == "A")
    type_b = sum(1 for t in tcs if t[7] == "B")

    apply_header_cell(ws, 5, 1, "Passed", bg=DARK_BLUE, font_bold=True, span_cols=2)
    apply_header_cell(ws, 5, 3, "Failed", bg=DARK_BLUE, font_bold=True, span_cols=2)
    apply_header_cell(ws, 5, 5, "Untested", bg=DARK_BLUE, font_bold=True, span_cols=2)
    apply_header_cell(ws, 5, 7, "N/A", bg=DARK_BLUE, font_bold=True)
    apply_header_cell(ws, 5, 8, "A", bg=DARK_BLUE, font_bold=True)
    apply_header_cell(ws, 5, 9, "B", bg=DARK_BLUE, font_bold=True)
    apply_header_cell(ws, 5, 10, "Total Test Cases", bg=DARK_BLUE, font_bold=True)

    # ======== ROW 6: Summary values ========
    apply_data_cell(ws, 6, 1, n_pass, bg=WHITE, bold=True); ws.merge_cells(f"A6:B6")
    apply_data_cell(ws, 6, 3, n_fail, bg=WHITE, bold=True); ws.merge_cells(f"C6:D6")
    apply_data_cell(ws, 6, 5, 0,      bg=WHITE, bold=True); ws.merge_cells(f"E6:F6")
    apply_data_cell(ws, 6, 7, type_n, bg=WHITE, bold=True)
    apply_data_cell(ws, 6, 8, type_a, bg=WHITE, bold=True)
    apply_data_cell(ws, 6, 9, type_b, bg=WHITE, bold=True)
    apply_data_cell(ws, 6, 10, n_tcs, bg=WHITE, bold=True)

    # ======== ROW 7: Empty separator ========
    for c in range(1, 11):
        ws.cell(row=7, column=c).border = thin_border()

    # ======== ROW 8: Column header ========
    headers = ["", "Condition", "Precondition", "Input", "Expected Return",
               "Exception", "Log Message", "Type", "P/F", "Exec Date"]
    for col, h in enumerate(headers, 1):
        apply_header_cell(ws, 8, col, h, bg=DARK_BLUE, font_bold=True)

    # UTCID row labels
    ws.merge_cells("A8:A9")
    apply_header_cell(ws, 8, 1, "", bg=DARK_BLUE)

    # ======== ROW 9: UTCID header ========
    for col, tc in enumerate(tcs, 2):
        pass  # single row layout; UTCIDs go in each data row

    # ======== DATA ROWS (starting from row 9) ========
    row = 9
    for tc in tcs:
        utcid, desc, precond, inputs, ret, exc, log, typ, pf, date = tc

        # Alternate row color
        bg = "EEF2FF" if row % 2 == 0 else WHITE

        apply_data_cell(ws, row, 1, utcid, bg=HEADER_BLUE, bold=True, color=WHITE)
        apply_data_cell(ws, row, 2, desc, bg=bg, align="left", wrap=True)
        apply_data_cell(ws, row, 3, precond, bg=bg, align="left", wrap=True)
        apply_data_cell(ws, row, 4, inputs, bg=bg, align="left", wrap=True)
        apply_data_cell(ws, row, 5, ret, bg=bg, align="left", wrap=True)
        apply_data_cell(ws, row, 6, exc, bg=bg, align="left", wrap=True)
        apply_data_cell(ws, row, 7, log, bg=bg, align="left", wrap=True)

        # Type with color coding
        type_bg = {"N": "C6EFCE", "A": "FFEB9C", "B": "BDD7EE"}.get(typ, WHITE)
        apply_data_cell(ws, row, 8, typ, bg=type_bg, bold=True)

        # P/F with color coding
        pf_bg = "C6EFCE" if pf == "P" else "FFC7CE"
        apply_data_cell(ws, row, 9, pf, bg=pf_bg, bold=True)

        apply_data_cell(ws, row, 10, date, bg=bg)
        ws.row_dimensions[row].height = 45
        row += 1

    # ======== LEGEND ========
    ws.row_dimensions[row].height = 15
    row += 1
    ws.merge_cells(f"A{row}:J{row}")
    legend = ws.cell(row=row, column=1,
        value="Type: N = Normal (Bình thường)   |   A = Abnormal (Bất thường)   |   B = Boundary (Biên giới)")
    legend.font = Font(name="Calibri", size=8, italic=True, color="595959")
    legend.alignment = left_align()

    ws.freeze_panes = "B9"


# ─── SUMMARY SHEET ────────────────────────────────────────────────────────────
def build_summary(wb, sheets_data):
    ws = wb.active
    ws.title = "SUMMARY"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1,
        value="📊 BÁO CÁO KIỂM THỬ - HOMEDECORESHOP")
    title.fill = hdr_fill(DARK_BLUE)
    title.font = Font(name="Calibri", color=WHITE, bold=True, size=16)
    title.alignment = center_align()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    sub = ws.cell(row=2, column=1,
        value=f"Tổng hợp Unit Test | Ngày tạo: Jun-25, 2025 | Người thực hiện: Tạ Đức Bảo")
    sub.font = Font(name="Calibri", color="595959", italic=True, size=10)
    sub.alignment = center_align()
    ws.row_dimensions[2].height = 20

    # Header row 4
    hdrs = ["STT", "Chức năng", "Jira", "Tổng TC", "Passed", "Failed", "N", "A", "B"]
    widths = [5, 35, 10, 10, 10, 10, 8, 8, 8]
    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        apply_header_cell(ws, 4, i, h, bg=DARK_BLUE, font_bold=True)

    jiras = ["HOM-3", "HOM-2", "HOM-4", "HOM-6", "HOM-7",
             "HOM-8", "HOM-9", "HOM-10", "HOM-11", "HOM-12", "HOM-13", "HOM-14", "HOM-15"]

    row = 5
    total_tc = total_pass = total_fail = 0
    for idx, sd in enumerate(sheets_data):
        _, func_name, _, _, jira, tcs = sd
        n = len(tcs)
        p = sum(1 for t in tcs if t[8] == "P")
        f = n - p
        type_n = sum(1 for t in tcs if t[7] == "N")
        type_a = sum(1 for t in tcs if t[7] == "A")
        type_b = sum(1 for t in tcs if t[7] == "B")

        bg = "EEF2FF" if row % 2 == 0 else WHITE
        apply_data_cell(ws, row, 1, idx + 1, bg=bg, bold=True)
        apply_data_cell(ws, row, 2, func_name, bg=bg, align="left")
        apply_data_cell(ws, row, 3, jira, bg=bg)
        apply_data_cell(ws, row, 4, n, bg=bg, bold=True)
        apply_data_cell(ws, row, 5, p, bg="C6EFCE", bold=True)
        apply_data_cell(ws, row, 6, f, bg="FFC7CE" if f > 0 else "C6EFCE", bold=True)
        apply_data_cell(ws, row, 7, type_n, bg="C6EFCE")
        apply_data_cell(ws, row, 8, type_a, bg="FFEB9C")
        apply_data_cell(ws, row, 9, type_b, bg="BDD7EE")

        total_tc += n; total_pass += p; total_fail += f
        row += 1

    # Total row
    ws.row_dimensions[row].height = 20
    apply_data_cell(ws, row, 1, "", bg=DARK_BLUE)
    apply_data_cell(ws, row, 2, "TỔNG CỘNG", bg=DARK_BLUE, bold=True, color=WHITE)
    apply_data_cell(ws, row, 3, "", bg=DARK_BLUE)
    apply_data_cell(ws, row, 4, total_tc, bg=DARK_BLUE, bold=True, color=WHITE)
    apply_data_cell(ws, row, 5, total_pass, bg=DARK_BLUE, bold=True, color=WHITE)
    apply_data_cell(ws, row, 6, total_fail, bg=DARK_BLUE, bold=True, color=WHITE)
    apply_data_cell(ws, row, 7, "", bg=DARK_BLUE)
    apply_data_cell(ws, row, 8, "", bg=DARK_BLUE)
    apply_data_cell(ws, row, 9, "", bg=DARK_BLUE)

    ws.freeze_panes = "A5"


build_summary(wb, SHEETS)
for sd in SHEETS:
    build_sheet(wb, sd)

out_path = r"c:\Users\LENOVO\Downloads\GIT\KCPM\TestCase_HomeDecorShop.xlsx"
wb.save(out_path)
print(f"✅ Đã tạo file Excel thành công: {out_path}")
